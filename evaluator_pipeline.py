import os
import sys
import math
import types
import importlib.util
from typing import Optional, Tuple, List, Dict

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def _ensure_streamlit_stub_if_missing() -> None:
    """Provide a minimal stub for streamlit to allow importing the main module in non-UI runs."""
    try:
        import streamlit  # noqa: F401
        return
    except Exception:
        pass

    try:
        import contextlib
    except Exception:
        contextlib = None

    st_module = types.ModuleType("streamlit")

    def cache_resource(show_spinner: bool = False):
        def decorator(fn):
            return fn
        return decorator

    st_module.cache_resource = cache_resource
    st_module.set_page_config = lambda *args, **kwargs: None
    st_module.title = lambda *args, **kwargs: None
    st_module.caption = lambda *args, **kwargs: None
    st_module.write = lambda *args, **kwargs: None
    st_module.warning = lambda *args, **kwargs: None
    st_module.text_input = lambda *args, **kwargs: ""
    st_module.slider = lambda *args, **kwargs: 25
    st_module.button = lambda *args, **kwargs: False
    if contextlib is not None:
        st_module.spinner = lambda *args, **kwargs: contextlib.nullcontext()
    else:
        class _DummyCtx:
            def __enter__(self):
                return None
            def __exit__(self, exc_type, exc, tb):
                return False
        st_module.spinner = lambda *args, **kwargs: _DummyCtx()

    sys.modules["streamlit"] = st_module


def load_hr_hub_module(project_root: str):
    """Load the main H&R Hub module by path so we can reuse its functions without editing it."""
    _ensure_streamlit_stub_if_missing()
    main_path = os.path.join(project_root, "H&R Hub.py")
    spec = importlib.util.spec_from_file_location("hr_hub_module", main_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {main_path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)  # type: ignore[attr-defined]
    return mod


def build_search_index(mod) -> Tuple[list, object, object]:
    """Load documents and build the vector index using the existing functions."""
    project_root = os.path.dirname(os.path.abspath(__file__))
    chunks = mod.load_corpus(project_root)
    vectorizer, matrix = mod.build_index(chunks)
    return chunks, vectorizer, matrix


def _strip_accents(s: str) -> str:
    try:
        import unicodedata
        nfkd = unicodedata.normalize("NFKD", s)
        return "".join(c for c in nfkd if not unicodedata.combining(c))
    except Exception:
        return s


def _normalize_header_text(s: str) -> str:
    import re
    s = (s or "").strip().lower()
    s = _strip_accents(s)
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _find_header_column_index(ws, header_name: str) -> Optional[int]:
    """Find 1-based column index by matching likely header variants within the first 3 rows."""
    target_norm = _normalize_header_text(header_name)
    variants = {target_norm}
    # Common variants for Questions and Ideal Answer
    if target_norm in {"questions", "question"}:
        variants.update({"question", "questions", "pregunta", "preguntas"})
    if target_norm in {"ideal answer", "ideal answer (ground truth)"}:
        variants.update({"ideal answer", "ground truth", "respuesta ideal"})

    for header_row in range(1, min(3, ws.max_row) + 1):
        exact_map = {}
        norm_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if isinstance(val, str):
                exact = val.strip()
                exact_map[exact] = col_idx
                norm = _normalize_header_text(exact)
                norm_map.setdefault(norm, col_idx)
        # direct exact
        if header_name in exact_map:
            return exact_map[header_name]
        # normalized equality or prefix
        if target_norm in norm_map:
            return norm_map[target_norm]
        for v in variants:
            if v in norm_map:
                return norm_map[v]
        # try prefix-based match (e.g., "questions (en)")
        for norm_val, idx in norm_map.items():
            if any(norm_val.startswith(v) for v in variants):
                return idx
    return None


def _ensure_chatbot_answer_column(ws, questions_col: int) -> int:
    """Ensure there is a 'Chatbot Answer' column positioned appropriately.
    If an 'Ideal Answer' exists, rename it to 'Chatbot Answer'. Otherwise, insert a new column right of 'Questions'.
    Returns the 1-based column index of the Chatbot Answer column.
    """
    ideal_idx = _find_header_column_index(ws, "Ideal Answer")
    chatbot_idx = _find_header_column_index(ws, "Chatbot Answer")
    if chatbot_idx is not None:
        return chatbot_idx
    if ideal_idx is not None:
        ws.cell(row=1, column=ideal_idx).value = "Chatbot Answer"
        return ideal_idx
    # Insert to the immediate right of Questions
    insert_at = questions_col + 1
    ws.insert_cols(insert_at)
    ws.cell(row=1, column=insert_at).value = "Chatbot Answer"
    # Give the new column a reasonable width for readability if none is set
    col_letter = get_column_letter(insert_at)
    if ws.column_dimensions[col_letter].width is None:
        ws.column_dimensions[col_letter].width = 60
    return insert_at


def _estimate_row_height_for_text(text: str, col_width_chars: float) -> float:
    if not text:
        return 15.0
    # Rough estimate: characters per line ~= column width; add padding
    chars_per_line = max(1.0, (col_width_chars or 50.0) * 1.0)
    lines = max(1, int(math.ceil(len(text) / chars_per_line)))
    # 15 pt per line is a common readable height, plus a bit of padding
    return float(min(15 * lines + 8, 140))


def _normalize_unavailable_to_english(answer: str) -> str:
    if not isinstance(answer, str):
        return ""
    trimmed = answer.strip()
    if trimmed.startswith("No se encuentra en la información disponible"):
        return "I cannot find information in the provided chunks to answer this."
    return trimmed


def generate_answer_for_question(query: str, mod, vectorizer, matrix, chunks, top_k: int = 25) -> str:
    def attempt_once() -> Optional[str]:
        ranked = mod.rank_chunks(query, vectorizer, matrix, chunks, top_k=top_k)
        ai_answer = mod.call_openai_generate(query, ranked, max_sentences=5)
        if ai_answer is None or (isinstance(ai_answer, str) and ai_answer.strip() == ""):
            fallback, _ = mod.compose_answer(query, ranked)
            return _normalize_unavailable_to_english(fallback)
        return str(ai_answer).strip()

    try:
        ans = attempt_once()
        if ans is not None and len(ans.strip()) > 0:
            return ans
        # retry once
        ans = attempt_once()
        if ans is not None and len(ans.strip()) > 0:
            return ans
    except Exception:
        pass
    return "Error al generar respuesta"


def create_chatbot_answers_excel(mod) -> str:
    """Read Evaluator questions and write Chatbot Answers.xlsx, preserving formatting where possible."""
    project_root = os.path.dirname(os.path.abspath(__file__))
    evaluator_path = os.path.join(project_root, "EVALUATOR", "Evaluator questions.xlsx")
    output_path = os.path.join(project_root, "EVALUATOR", "Chatbot Answers.xlsx")

    wb = load_workbook(evaluator_path)
    ws = wb.worksheets[0]

    q_col = _find_header_column_index(ws, "Questions")
    if not q_col:
        raise RuntimeError("Header 'Questions' not found in Evaluator questions.xlsx (first sheet)")

    chatbot_col = _ensure_chatbot_answer_column(ws, q_col)

    # Build index once
    chunks, vectorizer, matrix = build_search_index(mod)

    # Ensure header alignment for readability
    ws.cell(row=1, column=chatbot_col).alignment = Alignment(wrap_text=True, vertical="top")

    col_letter = get_column_letter(chatbot_col)
    col_width = ws.column_dimensions[col_letter].width or 50.0

    total = 0
    for row in range(2, ws.max_row + 1):
        q_val = ws.cell(row=row, column=q_col).value
        question = (str(q_val).strip() if q_val is not None else "")
        if not question:
            continue  # ignore empty rows
        total += 1
        answer = generate_answer_for_question(question, mod, vectorizer, matrix, chunks, top_k=25)
        cell = ws.cell(row=row, column=chatbot_col)
        cell.value = answer
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Adjust row height for readability
        ws.row_dimensions[row].height = _estimate_row_height_for_text(answer, float(col_width))

    # Save as a new workbook file
    wb.save(output_path)
    return output_path


def _tokenize_for_keywords(text: str) -> List[str]:
    if not text:
        return []
    import re
    tokens = [t.lower() for t in re.split(r"\W+", text) if t]
    # simple keyword heuristic: length >= 4
    return [t for t in tokens if len(t) >= 4]


def _tfidf_cosine_similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    try:
        from sklearn.feature_extraction.text import TfidfVectorizer
        from sklearn.metrics.pairwise import cosine_similarity
        vec = TfidfVectorizer(ngram_range=(1, 2), min_df=1)
        mat = vec.fit_transform([a, b])
        sim = float(cosine_similarity(mat[0:1], mat[1:2])[0][0])
        if sim != sim or sim < 0:  # nan or negative
            return 0.0
        return max(0.0, min(1.0, sim))
    except Exception:
        # Fallback: simple token overlap as proxy
        set_a = set(_tokenize_for_keywords(a))
        set_b = set(_tokenize_for_keywords(b))
        if not set_a or not set_b:
            return 0.0
        inter = len(set_a & set_b)
        denom = max(1, len(set_a | set_b))
        return inter / float(denom)


def _keyword_overlap_percent(a: str, b: str) -> float:
    set_a = set(_tokenize_for_keywords(a))
    set_b = set(_tokenize_for_keywords(b))
    if not set_a and not set_b:
        return 0.0
    inter = len(set_a & set_b)
    union = len(set_a | set_b) or 1
    return (inter / float(union)) * 100.0


def _length_diff_percent(a: str, b: str) -> float:
    la = len(a or "")
    lb = len(b or "")
    if la == 0 and lb == 0:
        return 0.0
    denom = max(1, la)
    return (abs(lb - la) / float(denom)) * 100.0


def _build_chatbot_mapping(ws_chat) -> Dict[str, List[str]]:
    q_col = _find_header_column_index(ws_chat, "Questions")
    a_col = _find_header_column_index(ws_chat, "Chatbot Answer")
    mapping: Dict[str, List[str]] = {}
    if not q_col or not a_col:
        return mapping
    for row in range(2, ws_chat.max_row + 1):
        q_val = ws_chat.cell(row=row, column=q_col).value
        a_val = ws_chat.cell(row=row, column=a_col).value
        question = (str(q_val).strip() if q_val is not None else "")
        answer = (str(a_val).strip() if a_val is not None else "")
        if not question:
            continue
        mapping.setdefault(question, []).append(answer)
    return mapping


def _strip_sources_section(text: str) -> str:
    if not text:
        return ""
    t = str(text).strip()
    # Remove everything after a line that starts with 'Sources:' (case-insensitive)
    lines = t.splitlines()
    out_lines: List[str] = []
    for ln in lines:
        if ln.strip().lower().startswith("sources:"):
            break
        out_lines.append(ln)
    return "\n".join(out_lines).strip()


def _evaluate_fidelity_openai(question: str, ideal: str, answer: str) -> Tuple[float, str]:
    """Use OpenAI to judge semantic fidelity between ideal and answer.
    Returns (fidelity_percent, notes). Falls back to TF-IDF if API unavailable.
    """
    ideal = (ideal or "").strip()
    answer = (answer or "").strip()
    if not answer:
        return 0.0, "Missing chatbot answer"

    # Strip trailing sources list to focus on content
    answer_core = _strip_sources_section(answer)

    try:
        from openai import OpenAI  # type: ignore
        api_key = os.environ.get("OPENAI_API_KEY", "").strip()
        if not api_key:
            raise RuntimeError("Missing OPENAI_API_KEY")
        client = OpenAI(api_key=api_key)

        system_msg = (
            "You are a meticulous evaluator. Compare a Ground Truth answer to a Chatbot answer for the same question. "
            "Judge SEMANTIC fidelity (meaning), not exact wording. Output STRICT JSON with keys: fidelity (0-100, number), notes (short string)."
        )
        user_msg = (
            "Question:\n" + question.strip() +
            "\n\nGround Truth:\n" + ideal +
            "\n\nChatbot Answer:\n" + answer_core +
            "\n\nRespond ONLY with JSON: {\"fidelity\": <number 0-100>, \"notes\": \"<short>\"}"
        )

        def _attempt() -> Optional[Tuple[float, str]]:
            try:
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {"role": "system", "content": system_msg},
                        {"role": "user", "content": user_msg},
                    ],
                    temperature=0.0,
                    max_tokens=100,
                )
                txt = (resp.choices[0].message.content or "").strip()
            except Exception:
                try:
                    resp2 = client.responses.create(
                        model="gpt-4o-mini",
                        input=[
                            {"role": "system", "content": system_msg},
                            {"role": "user", "content": user_msg},
                        ],
                        temperature=0.0,
                        max_output_tokens=100,
                    )
                    if hasattr(resp2, "output") and resp2.output and hasattr(resp2.output[0], "content"):
                        parts = resp2.output[0].content
                        if parts and hasattr(parts[0], "text"):
                            txt = (parts[0].text or "").strip()
                        else:
                            txt = ""
                    else:
                        txt = ""
                except Exception:
                    return None
            import json, re
            data = None
            try:
                data = json.loads(txt)
            except Exception:
                # Try to extract a number 0-100
                m = re.search(r"(\d+(?:\.\d+)?)", txt)
                if m:
                    try:
                        val = float(m.group(1))
                        return max(0.0, min(100.0, val)), ""
                    except Exception:
                        pass
                return None
            if not isinstance(data, dict) or "fidelity" not in data:
                return None
            try:
                val = float(data.get("fidelity", 0))
            except Exception:
                val = 0.0
            notes = str(data.get("notes", "")).strip()
            return max(0.0, min(100.0, val)), notes

        out = _attempt()
        if out is None:
            out = _attempt()
        if out is not None:
            return out
    except Exception:
        # Fall back
        pass

    # Fallback to tf-idf similarity
    sim = _tfidf_cosine_similarity(ideal, answer_core)
    return sim * 100.0, "TF-IDF fallback"


def create_metrics_excel() -> Tuple[str, int, int]:
    project_root = os.path.dirname(os.path.abspath(__file__))
    evaluator_path = os.path.join(project_root, "EVALUATOR", "Evaluator questions.xlsx")
    chatbot_path = os.path.join(project_root, "EVALUATOR", "Chatbot Answers.xlsx")
    output_path = os.path.join(project_root, "EVALUATOR", "Metrics H&R Hub Chatbot.xlsx")

    wb_eval = load_workbook(evaluator_path)
    ws_eval = wb_eval.worksheets[0]

    wb_chat = load_workbook(chatbot_path)
    ws_chat = wb_chat.worksheets[0]

    # Build mapping from question to sequential answers list
    q_to_answers = _build_chatbot_mapping(ws_chat)
    q_occurrence_counter: Dict[str, int] = {}

    # Determine where to place new columns (append at the end to avoid disturbing existing columns)
    start_col = ws_eval.max_column + 1
    headers = [
        "Ideal Answer (Ground Truth)",
        "Chatbot Answer",
        "Fidelity %",
        "High Fidelity",
        "Keyword Overlap %",
        "Length Difference %",
        "Notes",
    ]
    for i, h in enumerate(headers):
        ws_eval.cell(row=1, column=start_col + i).value = h

    q_col_eval = _find_header_column_index(ws_eval, "Questions")
    ideal_col_eval = _find_header_column_index(ws_eval, "Ideal Answer")

    total_questions = 0
    high_fidelity_count = 0

    for row in range(2, ws_eval.max_row + 1):
        q_val = ws_eval.cell(row=row, column=q_col_eval).value if q_col_eval else None
        question = (str(q_val).strip() if q_val is not None else "")
        if not question:
            continue
        total_questions += 1

        ideal_text = ""
        if ideal_col_eval:
            iv = ws_eval.cell(row=row, column=ideal_col_eval).value
            ideal_text = str(iv).strip() if iv is not None else ""

        # Fill Ideal Answer (Ground Truth)
        ws_eval.cell(row=1, column=start_col).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col).value = ideal_text
        ws_eval.cell(row=row, column=start_col).alignment = Alignment(wrap_text=True, vertical="top")

        # Fetch Chatbot Answer by occurrence
        occ = q_occurrence_counter.get(question, 0)
        answers_list = q_to_answers.get(question, [])
        chatbot_text = answers_list[occ] if occ < len(answers_list) else ""
        q_occurrence_counter[question] = occ + 1

        ws_eval.cell(row=1, column=start_col + 1).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 1).value = chatbot_text
        ws_eval.cell(row=row, column=start_col + 1).alignment = Alignment(wrap_text=True, vertical="top")

        # Metrics
        fidelity, ai_notes = _evaluate_fidelity_openai(question, ideal_text, chatbot_text)
        fidelity = max(0.0, min(100.0, fidelity))
        keyword_overlap = _keyword_overlap_percent(ideal_text, chatbot_text)
        length_diff = _length_diff_percent(ideal_text, chatbot_text)
        high_fidelity = fidelity > 85.0

        ws_eval.cell(row=1, column=start_col + 2).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 2).value = round(fidelity, 1)
        ws_eval.cell(row=1, column=start_col + 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 3).value = True if high_fidelity else False
        ws_eval.cell(row=1, column=start_col + 4).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 4).value = round(keyword_overlap, 1)
        ws_eval.cell(row=1, column=start_col + 5).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 5).value = round(length_diff, 1)

        note = ai_notes or ("Missing chatbot answer" if not chatbot_text else ("Low semantic similarity" if fidelity < 20.0 else ""))
        ws_eval.cell(row=1, column=start_col + 6).alignment = Alignment(wrap_text=True, vertical="top")
        ws_eval.cell(row=row, column=start_col + 6).value = note
        ws_eval.cell(row=row, column=start_col + 6).alignment = Alignment(wrap_text=True, vertical="top")

        if high_fidelity:
            high_fidelity_count += 1

    # Save metrics workbook
    wb_eval.save(output_path)
    return output_path, total_questions, high_fidelity_count


def main() -> None:
    project_root = os.path.dirname(os.path.abspath(__file__))
    mod = load_hr_hub_module(project_root)

    answers_path = create_chatbot_answers_excel(mod)
    print(f"Created: {answers_path}")

    metrics_path, total, high = create_metrics_excel()
    print(f"Created: {metrics_path}")
    print(f"Summary — Total questions: {total}; High fidelity (>85%): {high}")


if __name__ == "__main__":
    main()


