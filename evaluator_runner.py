# -*- coding: utf-8 -*-
"""
evaluator_runner.py — Evaluación LLM (único evaluador) robusta y lista para producción

Uso básico:
  python evaluator_runner.py
  python evaluator_runner.py metrics_llm

Parámetros útiles (opcionales):
  --model gpt-4o-mini       # o el que prefieras (por defecto lee OPENAI_EVAL_MODEL o gpt-4o-mini)
  --workers 4               # paraleliza llamadas al LLM (escritura en Excel sigue siendo single-thread)
  --fuzzy-threshold 0.92    # umbral de fuzzy matching (0–1)
  --no-fuzzy                # desactiva fuzzy matching (solo exacto normalizado)
  --evaluator-dir ./EVALUATOR
  --max-tokens 300
  --temperature 0.0

Requisitos:
  - openpyxl
  - openai>=1.0.0
  - Python 3.9+

Variables:
  - OPENAI_API_KEY (obligatoria)
  - OPENAI_EVAL_MODEL (opcional; por defecto: gpt-4o-mini)

Entradas esperadas (en ./EVALUATOR):
  - Evaluator questions.xlsx   (hoja con columna 'Questions' y, si existe, 'Ideal Answer')
  - Chatbot Answers.xlsx  o Chatbot Answers_*.xlsx  (con 'Questions' y 'Chatbot Answer')

Salida:
  - EVALUATOR/Metrics H&R Hub Chatbot (LLM)_{timestamp}.xlsx
  - EVALUATOR/.llm_cache.json (cache local para ahorrar tokens en re-ejecuciones)
"""
import argparse
import datetime
import hashlib
import json
import logging
import os
import re
import sys
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------- Logging ---------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("evaluator_runner")

# ---------------------------- Headers (multilenguaje) ------------------
QUESTION_HEADERS = ["Questions", "Question", "Pregunta", "Preguntas", "Pregunta(s)"]
IDEAL_HEADERS = [
    "Ideal Answer",
    "Ideal Answer (Ground Truth)",
    "Ground Truth",
    "Respuesta ideal",
    "Respuesta Ideal",
]
CHATBOT_HEADERS = ["Chatbot Answer", "Respuesta Chatbot", "Chatbot"]

# ---------------------------- Utils de texto/Excel ---------------------
def normalize_ws(x: Optional[str]) -> str:
    if x is None:
        return ""
    return " ".join(str(x).split())

def safe_lower(s: Optional[str]) -> str:
    return normalize_ws(s).lower()

def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

def normalize_question_key(s: str) -> str:
    t = strip_accents(normalize_ws(s)).lower()
    t = re.sub(r"\s+", " ", t)
    # Elimina todo excepto letras/números para que "¿Qué es...?" == "que es"
    t = re.sub(r"[^a-z0-9]+", "", t)
    return t

def find_header_col(ws, wanted_names: List[str]) -> Optional[int]:
    """Devuelve índice de columna (1-based) cuyo header coincide con alguno de wanted_names (case-insensitive)."""
    headers = {}
    header_list = []
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=col).value
        hdr = safe_lower(raw)
        headers[hdr] = col
        header_list.append((col, hdr))
    # exacto
    for name in wanted_names:
        nm = safe_lower(name)
        if nm in headers:
            return headers[nm]
    # contains
    for col, hdr in header_list:
        if not hdr:
            continue
        for name in wanted_names:
            nm = safe_lower(name)
            if nm and (nm in hdr or hdr in nm):
                return col
    return None

def select_first_sheet_with_header(wb, wanted_names: List[str]):
    for ws in wb.worksheets:
        if find_header_col(ws, wanted_names) is not None:
            return ws
    return wb.worksheets[0] if wb.worksheets else None

def set_col_wrap(ws, col_idx: int, vertical: str = "top"):
    for r in range(1, ws.max_row + 1):
        c = ws.cell(row=r, column=col_idx)
        cur = c.alignment or Alignment()
        c.alignment = Alignment(wrap_text=True, horizontal=cur.horizontal, vertical=vertical)

def estimate_row_height(text: str, approx_chars_per_line: float) -> float:
    if not text:
        return 15.0
    lines = text.count("\n") + max(1, int((len(text) + approx_chars_per_line - 1) // max(1.0, approx_chars_per_line)))
    height = 15.0 + (lines - 1) * 12.0
    return min(height, 220.0)

def column_width(ws, col_idx: int) -> float:
    letter = get_column_letter(col_idx)
    cd = ws.column_dimensions.get(letter)
    return float(cd.width) if (cd and cd.width) else 10.0

def set_column_width(ws, col_idx: int, width: float):
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = width

def insert_blank_column(ws, insert_at: int, copy_width_from: Optional[int] = None) -> None:
    ws.insert_cols(insert_at, amount=1)
    if copy_width_from:
        set_column_width(ws, insert_at, column_width(ws, copy_width_from))

# ---------------------------- Carga de filas ---------------------------
def load_rows(ws, q_col: int, ia_col: Optional[int]) -> List[Tuple[int, str, str]]:
    out = []
    for r in range(2, ws.max_row + 1):
        q = normalize_ws(ws.cell(row=r, column=q_col).value)
        if not q:
            continue
        ideal = normalize_ws(ws.cell(row=r, column=ia_col).value) if ia_col else ""
        out.append((r, q, ideal))
    return out

def load_chatbot_rows(ws, q_col: int, ch_col: int) -> List[Tuple[int, str, str]]:
    out = []
    for r in range(2, ws.max_row + 1):
        q = normalize_ws(ws.cell(row=r, column=q_col).value)
        if not q:
            continue
        ch = normalize_ws(ws.cell(row=r, column=ch_col).value)
        out.append((r, q, ch))
    return out

# ---------------------------- Emparejamiento ---------------------------
def match_by_question_smart(eval_rows, chat_rows, fuzzy_threshold: float = 0.92, enable_fuzzy: bool = True):
    """
    Empareja por clave normalizada; si falla, intenta emparejamiento borroso (difflib).
    Mantiene el orden para duplicados (colas por clave).
    """
    from collections import defaultdict, deque
    from difflib import SequenceMatcher

    buckets = defaultdict(deque)
    for r, q, ch in chat_rows:
        buckets[normalize_question_key(q)].append((r, q, ch))

    pairings, pending = [], []
    # 1) Exacto por clave normalizada
    for r_eval, q, ideal in eval_rows:
        k = normalize_question_key(q)
        if buckets[k]:
            r_chat, _q_orig, ch = buckets[k].popleft()
            pairings.append((r_eval, q, ideal, r_chat, ch))
        else:
            pending.append((r_eval, q, ideal))

    if not enable_fuzzy or not pending:
        # Añadir no emparejadas con chatbot vacío
        for r_eval, q, ideal in pending:
            pairings.append((r_eval, q, ideal, None, ""))
        return pairings

    # 2) Fuzzy para lo pendiente
    remaining = []
    for k, dq in buckets.items():
        while dq:
            r_chat, q_orig, ch = dq.popleft()
            remaining.append((k, r_chat, q_orig, ch))

    used = set()
    for r_eval, q, ideal in pending:
        k = normalize_question_key(q)
        best = None
        for k2, r_chat, q_orig, ch in remaining:
            if r_chat in used:
                continue
            ratio = SequenceMatcher(None, k, k2).ratio()
            if ratio >= fuzzy_threshold and (best is None or ratio > best[0]):
                best = (ratio, r_chat, ch)
        if best:
            _, r_chat, ch = best
            used.add(r_chat)
            pairings.append((r_eval, q, ideal, r_chat, ch))
        else:
            pairings.append((r_eval, q, ideal, None, ""))
    return pairings

# ---------------------------- Localiza Answers -------------------------
def _find_answers_workbook(evaluator_dir: Path) -> Path:
    fixed = evaluator_dir / "Chatbot Answers.xlsx"
    if fixed.exists():
        return fixed
    cands = sorted(evaluator_dir.glob("Chatbot Answers_*.xlsx"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if cands:
        return cands[0]
    raise FileNotFoundError(
        "No se encontró 'Chatbot Answers.xlsx' ni 'Chatbot Answers_*.xlsx' en EVALUATOR/."
    )

# ---------------------------- JSON robusto -----------------------------
def _strip_code_fences(payload: str) -> str:
    """
    Extrae de forma robusta el primer objeto JSON del payload, tolerando ```json fences
    y texto antes/después. Devuelve una cadena JSON (serializada) o el payload original.
    """
    s = (payload or "").strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?", "", s, flags=re.IGNORECASE).strip()
        s = re.sub(r"```$", "", s).strip()
    start = s.find("{")
    if start != -1:
        try:
            obj, _ = json.JSONDecoder().raw_decode(s[start:])
            return json.dumps(obj)
        except json.JSONDecodeError:
            pass
    return s

# ---------------------------- LLM Evaluator ----------------------------
def _cache_key(model: str, question: str, ideal: str, chatbot: str) -> str:
    h = hashlib.sha256()
    h.update(model.encode("utf-8"))
    h.update(b"\nQ:")
    h.update(question.encode("utf-8"))
    h.update(b"\nI:")
    h.update(ideal.encode("utf-8"))
    h.update(b"\nC:")
    h.update(chatbot.encode("utf-8"))
    return h.hexdigest()

def _normalize_unavailable_to_english(answer: str) -> str:
    if not isinstance(answer, str):
        return ""
    trimmed = answer.strip()
    if trimmed.startswith("No se encuentra en la información disponible"):
        return "I cannot find information in the provided chunks to answer this."
    return trimmed

def generate_answer_for_question(query: str, mod, vectorizer, matrix, chunks, top_k: int = 25) -> str:
    def attempt_once() -> Optional[str]:
        ranked = mod.rank_chunks(query, vectorizer, matrix, chunks, top_k=top_k)
        ai_answer, _ = mod.call_openai_generate(query, ranked, max_sentences=5)
        if ai_answer is None or (isinstance(ai_answer, str) and ai_answer.strip() == ""):
            fallback, _ = mod.compose_answer(query, ranked)
            return _normalize_unavailable_to_english(fallback)
        return str(ai_answer).strip()

    try:
        ans = attempt_once()
        if ans is not None and len(ans.strip()) > 0:
            return ans
        # retry once
        ans = attempt_once()
        if ans is not None and len(ans.strip()) > 0:
            return ans
    except Exception:
        pass
    return "Error al generar respuesta"

def call_openai_evaluator(
    question: str,
    ideal_answer: str,
    chatbot_answer: str,
    model: str,
    temperature: float = 0.0,
    max_tokens: int = 300,
    cache: Optional[Dict[str, Dict]] = None,
) -> Tuple[Optional[bool], Optional[int], str]:
    """
    Juzga si la respuesta del chatbot es suficientemente correcta frente a la ideal.
    Devuelve (sufficiently_correct(bool|None), score(int|None), notes(str)).
    Usa cache si está disponible.
    """
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        return None, None, "Missing OPENAI_API_KEY"

    ck = _cache_key(model, question, ideal_answer, chatbot_answer)
    if cache is not None and ck in cache:
        data = cache[ck]
        return data.get("sufficiently_correct"), data.get("score"), data.get("notes", "")

    system_msg = (
        "You are an expert evaluator for enterprise knowledge-grounded Q&A. "
        "Assess whether the chatbot answer is sufficiently correct compared to the IDEAL answer for the same question. "
        "Be fair and consider whether the provided answer goes in the same direction and alignment with the ideal answer; "
        "it doesn’t need to be perfect, just correct enough or close enough. "
        "Return ONLY strict JSON with keys: sufficiently_correct (boolean), score (integer 1-10), notes (short reason)."
    )
    user_msg = (
        "Evaluate the following answers to the same question.\n\n"
        f"Question:\n{question}\n\n"
        f"IDEAL ANSWER:\n{ideal_answer}\n\n"
        f"CHATBOT ANSWER:\n{chatbot_answer}\n\n"
        "Output format (JSON only): {\n"
        "  \"sufficiently_correct\": true|false,\n"
        "  \"score\": <integer 1-10>,\n"
        "  \"notes\": \"short reason\"\n"
        "}"
    )

    client = OpenAI(api_key=api_key)
    last_err = None
    for attempt in range(3):
        try:
            # Intento con respuesta JSON estricta
            try:
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "system", "content": system_msg},
                              {"role": "user", "content": user_msg}],
                    temperature=temperature,
                    max_tokens=max_tokens,
                    response_format={"type": "json_object"},
                )
                text = (resp.choices[0].message.content or "").strip()
            except Exception:
                # Fallback sin response_format
                resp = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "system", "content": system_msg},
                              {"role": "user", "content": user_msg}],
                    temperature=temperature,
                    max_tokens=max_tokens,
                )
                text = (resp.choices[0].message.content or "").strip()

            payload = _strip_code_fences(text)
            try:
                data = json.loads(payload)
            except Exception:
                start = payload.find("{")
                if start != -1:
                    try:
                        data, _ = json.JSONDecoder().raw_decode(payload[start:])
                    except Exception:
                        return None, None, f"Unparseable fields in response: {payload[:180]}"
                else:
                    return None, None, f"Unparseable fields in response: {payload[:180]}"

            sc = data.get("sufficiently_correct")
            score = data.get("score")
            notes = str(data.get("notes", ""))

            if isinstance(sc, bool) and isinstance(score, int):
                if cache is not None:
                    cache[ck] = {"sufficiently_correct": sc, "score": score, "notes": notes}
                return sc, max(1, min(10, score)), notes
            if isinstance(sc, bool) and isinstance(score, str) and score.isdigit():
                si = int(score)
                if cache is not None:
                    cache[ck] = {"sufficiently_correct": sc, "score": si, "notes": notes}
                return sc, max(1, min(10, si)), notes
            return None, None, f"Unparseable fields in response: {str(data)[:180]}"
        except Exception as e:
            last_err = e
            # backoff con jitter suave
            time.sleep((2 ** attempt) + (0.1 * attempt))
    return None, None, f"OpenAI error after retries: {last_err}"

# ---------------------------- Excel helpers ----------------------------
def _apply_conditional_formatting(ws, bool_col_idx: int, score_col_idx: int, max_row: int):
    """Colorea TRUE/ FALSE y gradiente 1–10 para score."""
    bool_letter = get_column_letter(bool_col_idx)
    score_letter = get_column_letter(score_col_idx)
    bool_range = f"{bool_letter}2:{bool_letter}{max_row}"
    score_range = f"{score_letter}2:{score_letter}{max_row}"

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

    ws.conditional_formatting.add(bool_range, CellIsRule(operator="equal", formula=["TRUE"], fill=green))
    ws.conditional_formatting.add(bool_range, CellIsRule(operator="equal", formula=["FALSE"], fill=red))

    ws.conditional_formatting.add(
        score_range,
        ColorScaleRule(
            start_type="min", start_color="F8696B",  # rojo
            mid_type="percentile", mid_value=50, mid_color="FFEB84",  # amarillo
            end_type="max", end_color="63BE7B",  # verde
        ),
    )

# ---------------------------- Métricas LLM -----------------------------
def build_metrics_workbook_llm(
    evaluator_dir: Path,
    model: str,
    workers: int = 1,
    fuzzy_threshold: float = 0.92,
    enable_fuzzy: bool = True,
    max_tokens: int = 300,
    temperature: float = 0.0,
    no_cache: bool = True,  # Default to True to disable cache by default
) -> Path:
    src_path = evaluator_dir / "Evaluator questions.xlsx"
    if not src_path.exists():
        raise FileNotFoundError(f"Missing template workbook: {src_path}")

    # Load H&R Hub module
    import types
    import importlib.util
    def _ensure_streamlit_stub_if_missing() -> None:
        try:
            import streamlit  # noqa: F401
            return
        except Exception:
            pass
        st_module = types.ModuleType("streamlit")
        def cache_resource(show_spinner: bool = False):
            def decorator(fn):
                return fn
            return decorator
        st_module.cache_resource = cache_resource
        st_module.set_page_config = lambda *args, **kwargs: None
        st_module.title = lambda *args, **kwargs: None
        st_module.caption = lambda *args, **kwargs: None
        st_module.write = lambda *args, **kwargs: None
        st_module.warning = lambda *args, **kwargs: None
        st_module.text_input = lambda *args, **kwargs: ""
        st_module.slider = lambda *args, **kwargs: 25
        st_module.button = lambda *args, **kwargs: False
        class _DummyCtx:
            def __enter__(self):
                return None
            def __exit__(self, exc_type, exc, tb):
                return False
        st_module.spinner = lambda *args, **kwargs: _DummyCtx()
        sys.modules["streamlit"] = st_module

    project_root = Path(__file__).resolve().parent
    main_path = project_root / "H&R Hub.py"
    spec = importlib.util.spec_from_file_location("hr_hub_module", main_path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load module from {main_path}")
    mod = importlib.util.module_from_spec(spec)
    _ensure_streamlit_stub_if_missing()
    spec.loader.exec_module(mod)

    # Build index
    chunks = mod.load_corpus(project_root)
    vectorizer, matrix = mod.build_index(chunks)

    # Carga hojas correctas por encabezados
    wb = load_workbook(filename=str(src_path))
    ws = select_first_sheet_with_header(wb, QUESTION_HEADERS)
    if ws is None:
        raise ValueError("No worksheet found with a 'Questions' header in Evaluator questions.")
    q_col = find_header_col(ws, QUESTION_HEADERS)
    ia_col = find_header_col(ws, IDEAL_HEADERS)
    if q_col is None:
        raise ValueError("Template workbook must contain a 'Questions' header in row 1.")

    # Extrae filas
    eval_rows = load_rows(ws, q_col, ia_col)

    # Generate chatbot answers on the fly
    chat_rows = []
    for r, q, ideal in eval_rows:
        if not q:
            continue
        ch = generate_answer_for_question(q, mod, vectorizer, matrix, chunks, top_k=200)
        chat_rows.append((r, q, ch))

    # Pairings: since generated directly, pair them
    pairings = []
    for (r_eval, q, ideal), (r_chat, _q, ch) in zip(eval_rows, chat_rows):
        pairings.append((r_eval, q, ideal, r_chat, ch))

    # Avisos (simplified since no unmatched)
    logger.info("Generated %d chatbot answers for %d questions.", len(chat_rows), len(eval_rows))

    # Inserta columna 'Chatbot Answer' (junto a Ideal si existe; si no, junto a Questions)
    dest_chat_col = (ia_col + 1) if ia_col else (q_col + 1)
    insert_blank_column(ws, dest_chat_col, copy_width_from=ia_col or q_col)
    ws.cell(row=1, column=dest_chat_col).value = "Chatbot Answer"
    set_col_wrap(ws, dest_chat_col)
    if column_width(ws, dest_chat_col) < 40.0:
        set_column_width(ws, dest_chat_col, 70.0)

    if ia_col is not None:
        ws.cell(row=1, column=ia_col).value = "Ideal Answer (Ground Truth)"
        set_col_wrap(ws, ia_col)

    # Columnas de métrica LLM
    labels = ["Sufficiently Correct", "Score (1–10)", "Evaluator Notes"]
    base = dest_chat_col + 1
    for i, lab in enumerate(labels):
        insert_blank_column(ws, base + i, copy_width_from=dest_chat_col)
        ws.cell(row=1, column=base + i).value = lab
        set_col_wrap(ws, base + i)
        if lab == "Evaluator Notes":
            set_column_width(ws, base + i, max(30.0, column_width(ws, dest_chat_col) * 0.5))

    # Alturas de fila y placeholders iniciales
    approx_cpl_chat = max(18.0, column_width(ws, dest_chat_col) * 1.1)
    approx_cpl_ideal = max(18.0, column_width(ws, ia_col) * 1.1) if ia_col else 50.0
    approx_cpl_q = max(18.0, column_width(ws, q_col) * 1.1)

    missing_ideal = 0
    missing_chat = 0
    for (r_eval, q, ideal, _r_chat, chatbot) in pairings:
        ws.cell(row=r_eval, column=dest_chat_col).value = chatbot
        if not normalize_ws(chatbot):
            missing_chat += 1
        if ia_col is not None and not normalize_ws(ideal):
            missing_ideal += 1
        h0 = estimate_row_height(q, approx_cpl_q)
        h1 = estimate_row_height(chatbot, approx_cpl_chat)
        h2 = estimate_row_height(ideal, approx_cpl_ideal)
        ws.row_dimensions[r_eval].height = max(h0, h1, h2)

    # Usabilidad: congelar cabecera y activar autofiltro
    try:
        ws.freeze_panes = ws["A2"]
        ws.auto_filter.ref = ws.dimensions
    except Exception:
        pass

    # ---------- Llamadas al LLM (con cache y paralelización opcional) ----------
    cache_path = evaluator_dir / ".llm_cache.json"
    cache: Dict[str, Dict] = {}
    if not no_cache and cache_path.exists():
        try:
            cache.update(json.loads(cache_path.read_text(encoding="utf-8")))
            logger.info("Cache LLM cargada (%d entradas).", len(cache))
        except Exception:
            logger.warning("No se pudo cargar la cache LLM, se reiniciará.")

    # Prep trabajo
    jobs = []
    for (r_eval, q, ideal, _r_chat, chatbot) in pairings:
        if not normalize_ws(chatbot):
            # sin chatbot → marcamos como falso/None con nota
            ws.cell(row=r_eval, column=base + 0).value = False
            ws.cell(row=r_eval, column=base + 1).value = None
            ws.cell(row=r_eval, column=base + 2).value = "Missing chatbot answer"
            continue
        if ia_col is not None and not normalize_ws(ideal):
            ws.cell(row=r_eval, column=base + 0).value = None
            ws.cell(row=r_eval, column=base + 1).value = None
            ws.cell(row=r_eval, column=base + 2).value = "Missing ideal answer"
            continue
        jobs.append((r_eval, q, ideal, chatbot))

    logger.info("Preguntas totales: %d | a evaluar con LLM: %d | sin ideal: %d | sin chatbot: %d",
                len(pairings), len(jobs), missing_ideal, missing_chat)

    def _eval_job(item):
        r_eval, q, ideal, chatbot = item
        sc, score, notes = call_openai_evaluator(
            q, ideal, chatbot, model=model, temperature=temperature, max_tokens=max_tokens, cache=(cache if not no_cache else None)  # Pass None if --no-cache
        )
        return r_eval, sc, score, notes

    results: Dict[int, Tuple[Optional[bool], Optional[int], str]] = {}
    if workers and workers > 1:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(_eval_job, job) for job in jobs]
            for i, fut in enumerate(as_completed(futs), 1):
                r_eval, sc, score, notes = fut.result()
                results[r_eval] = (sc, score, notes)
                if i % 10 == 0 or i == len(futs):
                    logger.info("Progreso LLM: %d/%d", i, len(futs))
    else:
        for i, job in enumerate(jobs, 1):
            r_eval, sc, score, notes = _eval_job(job)
            results[r_eval] = (sc, score, notes)
            if i % 10 == 0 or i == len(jobs):
                logger.info("Progreso LLM: %d/%d", i, len(jobs))

    # Aplicar resultados al Excel (solo desde el hilo principal)
    total = len(pairings)
    sufficient = 0
    sum_scores = 0
    count_scores = 0
    for (r_eval, _q, _ideal, _r_chat, _chatbot) in pairings:
        if r_eval in results:
            sc, score, notes = results[r_eval]
            if sc is True:
                sufficient += 1
            if score is not None:
                sum_scores += score
                count_scores += 1
            ws.cell(row=r_eval, column=base + 0).value = sc if sc is not None else None
            ws.cell(row=r_eval, column=base + 1).value = score if score is not None else None
            ws.cell(row=r_eval, column=base + 2).value = notes

    # Condicionales de color
    try:
        _apply_conditional_formatting(ws, base + 0, base + 1, ws.max_row)
    except Exception:
        pass

    # Hoja Summary (limpia y reescribe fila 1..)
    if "Summary" in wb.sheetnames:
        summary_ws = wb["Summary"]
        for row in summary_ws["A1:F10"]:
            for cell in row:
                cell.value = None
    else:
        summary_ws = wb.create_sheet("Summary")

    avg_score = (sum_scores / count_scores) if count_scores else 0.0
    summary_ws["A1"] = "Total Questions";              summary_ws["B1"] = total
    summary_ws["A2"] = "Sufficiently Correct";         summary_ws["B2"] = sufficient
    summary_ws["A3"] = "Sufficiently Correct %";       summary_ws["B3"] = round(100.0 * (sufficient / total if total else 0.0), 1)
    summary_ws["A4"] = "Average Score (1–10)";         summary_ws["B4"] = round(avg_score, 2)
    summary_ws["A6"] = "Model";                        summary_ws["B6"] = model
    summary_ws["A7"] = "Workers";                      summary_ws["B7"] = workers
    summary_ws["A8"] = "Fuzzy Threshold";              summary_ws["B8"] = fuzzy_threshold if enable_fuzzy else "Disabled"
    summary_ws["A9"] = "Run Timestamp";                summary_ws["B9"] = datetime.datetime.now().isoformat(timespec="seconds")

    # Guardar cache (skip if --no-cache)
    if not no_cache:
        try:
            cache_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
        except Exception:
            logger.warning("No se pudo escribir la cache LLM.")

    # Guardar archivo de salida
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = evaluator_dir / f"Metrics H&R Hub Chatbot (LLM)_{timestamp}.xlsx"
    try:
        wb.save(str(out_path))
        logger.info("Saved: %s", out_path)
        logger.info(
            "LLM Summary — Total: %s | Sufficient: %s (%.1f%%) | Avg Score: %.2f",
            total, sufficient, 100.0 * (sufficient / total if total else 0.0), avg_score
        )
    except Exception as e:
        alt = evaluator_dir / f"Metrics H&R Hub Chatbot (LLM)_{timestamp}_UNSAVED.xlsx"
        logger.error("Error al guardar %s; intento alternativo: %s | error=%s", out_path, alt, e)
        wb.save(str(alt))
        return alt
    return out_path

# ---------------------------- CLI / Main --------------------------------
def parse_args(argv: List[str]):
    parser = argparse.ArgumentParser(description="Evaluator Runner — metrics_llm")
    parser.add_argument("cmd", nargs="?", default="metrics_llm", choices=["metrics_llm"])
    parser.add_argument("--model", default=os.environ.get("OPENAI_EVAL_MODEL", "gpt-4o-mini"))
    parser.add_argument("--workers", type=int, default=1)
    parser.add_argument("--fuzzy-threshold", type=float, default=0.92)
    parser.add_argument("--no-fuzzy", action="store_true")
    parser.add_argument("--evaluator-dir", default=None)
    parser.add_argument("--max-tokens", type=int, default=300)
    parser.add_argument("--temperature", type=float, default=0.0)
    parser.add_argument("--cache", action="store_true", help="Enable LLM cache to speed up evaluations (default: disabled)")
    return parser.parse_args(argv[1:])

def main(argv: List[str]) -> None:
    args = parse_args(argv)
    project_root = Path(__file__).resolve().parent
    evaluator_dir = Path(args.evaluator_dir) if args.evaluator_dir else (project_root / "EVALUATOR")
    evaluator_dir.mkdir(parents=True, exist_ok=True)

    if args.cmd == "metrics_llm":
        path = build_metrics_workbook_llm(
            evaluator_dir=evaluator_dir,
            model=args.model,
            workers=max(1, int(args.workers)),
            fuzzy_threshold=float(args.fuzzy_threshold),
            enable_fuzzy=(not args.no_fuzzy),
            max_tokens=int(args.max_tokens),
            temperature=float(args.temperature),
            no_cache=not args.cache,  # Invert the flag: if --cache is set, use cache (no_cache=False)
        )
        logger.info("Done. Metrics (LLM): %s", path)
    else:
        logger.error("Unknown command: %s\nUse: python evaluator_runner.py metrics_llm", args.cmd)

if __name__ == "__main__":
    main(sys.argv)